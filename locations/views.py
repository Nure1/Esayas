from django.shortcuts import render
from django.http import HttpResponse
from rest_framework import generics 
from .models import Location 
from .serializers import LocationSerializer 
from openpyxl import Workbook
from rest_framework.views import APIView
from rest_framework.response import Response
from django.contrib.auth.models import User
from .serializers import UserSerializer
from django.contrib.auth import authenticate
from django.template import loader

class LocationListAPIView(generics.ListAPIView): 
    queryset = Location.objects.all() 
    serializer_class = LocationSerializer

class LocationUpdateAPIView(generics.UpdateAPIView):
    queryset = Location.objects.all()
    serializer_class = LocationSerializer

class LocationCreateAPIView(generics.CreateAPIView):
    queryset = Location.objects.all()
    serializer_class = LocationSerializer

class LocationDeleteAPIView(generics.DestroyAPIView):
    queryset = Location.objects.all()
    serializer_class = LocationSerializer

class UserLoginAPIView(APIView):
    #   Convert a user token into user data
	def get (self, request, format=None):
		if request.user.is_authenticated == False or request.user.is_active == False:
			return Response("Invalid Credentials", status=403)
		user = UserSerializer(request.user)
		print("Login Class login")
		return Response(user.data, status=200)
	def post(self, request, format=None):
		print("Login Class")
		user_obj = User.objects.filter(email=request.data['username']).first() or User.objects.filter(username=request.data['username']).first()
		if user_obj is not None:
			credentials={
				'username': user_obj.username,
				'password': request.data['password']
			}
		user= authenticate(**credentials)

		if user and user.is_active:
			user_serializer = UserSerializer(user)
			return Response({"user":user_serializer.data}, status=200)
		return Response("Invalid Credentials", status=403)

def export_to_excel(request):
    # Fetch data from the model
    locations = Location.objects.all()


    # Create a new workbook and add a worksheet
    workbook = Workbook()
    worksheet = workbook.active


    # Get all field names from the Location model
    field_names = [field.name for field in Location._meta.get_fields() if field.concrete]


    # Write column headers
    for col_num, field_name in enumerate(field_names, 1):
        worksheet.cell(row=1, column=col_num, value=field_name)


    # Write data rows
    for row_num, location in enumerate(locations, 2):
        for col_num, field_name in enumerate(field_names, 1):
            if field_name == 'user':
                user_value = getattr(location.user, 'username', '')
                worksheet.cell(row=row_num, column=col_num, value=user_value)
            else:
                field_value = getattr(location, field_name, '')
                worksheet.cell(row=row_num, column=col_num, value=field_value)


    # Set the response content type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=locations.xlsx'


    # Save the workbook to the response
    workbook.save(response)


    return response

def home(request):
    template = loader.get_template('index.html')
    return HttpResponse(template.render())

def about(request):
    template = loader.get_template('about.html')
    return HttpResponse(template.render())

def portfolio(request):
    template = loader.get_template('portfolio.html')
    return HttpResponse(template.render())

def service(request):
    template = loader.get_template('service.html')
    return HttpResponse(template.render())

def contact(request):
    template = loader.get_template('contact.html')
    return HttpResponse(template.render())

def login(request):
    template = loader.get_template('login.html')
    return HttpResponse(template.render())